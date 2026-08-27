# Importa el mÃ³dulo os para interactuar con el sistema operativo (rutas, archivos)
import os
import re
import datetime
import json
# Importa uuid para generar identificadores Ãºnicos universales (usado en radicados fÃ­sicos)
import uuid
# Importa funciones Ãºtiles de Django: renderizar templates, obtener objetos seguros y redireccionar
from django.shortcuts import render, get_object_or_404, redirect
# Importa el decorador que exige que un usuario estÃ© logueado para acceder a una vista
from django.contrib.auth.decorators import login_required
# Importa el modelo de Usuario por defecto de Django
from django.contrib.auth.models import User
# Importa el sistema de mensajes de Django para mostrar alertas (Ã©xito, error, etc.)
from django.contrib import messages
# Importa Q para hacer consultas complejas en la base de datos (ej. AND, OR)
from django.db.models import Q
# Importa timezone de Django para manejar fechas con conocimiento de la zona horaria
from django.utils import timezone

# Importa los modelos locales de la aplicaciÃ³n 'afiliados'
from .models import Carpeta, HistorialCarpeta, Profile
# Importa los formularios locales para validación y creación de datos
from .forms import CarpetaForm
# Importa los modelos de la aplicaciÃ³n 'documentos' relacionados con la correspondencia
from documentos.models import Documento, TipoDocumento
# Importa los modelos de la aplicaciÃ³n 'ubicacion' para el manejo fÃ­sico del archivo


# Funciones auxiliares para control de roles jerárquicos
def is_super(user):
    return user.is_superuser or (hasattr(user, 'profile') and user.profile.rol == 'SUPER')

def is_jefe(user):
    return is_super(user) or (hasattr(user, 'profile') and user.profile.rol == 'JEFE')

def is_aux_or_higher(user):
    return is_jefe(user) or (hasattr(user, 'profile') and user.profile.rol == 'AUX')

# API pública en formato JSON para consultas remotas del Gestor de Escritorio
from django.http import JsonResponse

def api_buscar_afiliado(request):
    try:
        query = request.GET.get('q', '').strip()
        if not query:
            return JsonResponse({'error': 'Parámetro q requerido'}, status=400)
        
        clean_q = re.sub(r"\D", "", query)
        qs = Carpeta.objects.filter(categoria='TRABAJADOR')
        matches = []
        if clean_q:
            # 1. Búsqueda exacta ultra-rápida por índice de cédula
            exact_match = qs.filter(identificacion=clean_q).first()
            if exact_match:
                matches = [exact_match]
                # Complementar con coincidencias adicionales si existen
                other_matches = list(qs.filter(Q(identificacion__icontains=clean_q) | Q(nombre__icontains=query)).exclude(id=exact_match.id)[:49])
                matches.extend(other_matches)
            else:
                matches = list(qs.filter(Q(identificacion__icontains=clean_q) | Q(nombre__icontains=query))[:50])
        else:
            matches = list(qs.filter(nombre__icontains=query)[:50])

            
        if not matches:
            return JsonResponse({'encontrado': False, 'resultados': [], 'mensaje': 'Expediente no encontrado'}, status=200)
            
        results = []
        for match in matches:
            fecha_str = ""
            if match.fecha:
                fecha_str = str(match.fecha)
            elif getattr(match, 'fecha_registro', None):
                fecha_str = match.fecha_registro.strftime("%Y-%m-%d")

            results.append({
                'Fecha': fecha_str,
                'Tipo Identificación': match.tipo_identificacion or "CC",
                'Cédula': match.identificacion,
                'Nombre': match.nombre or "SIN NOMBRE",
                'Módulo': str(match.modulo),
                'Estante': str(match.estante),
                'Bandeja': str(match.bandeja),
                'Cubículo': str(match.cubiculo),
                'Número de Carpeta': str(match.numero_carpeta),
                'Estado': match.estado or "ACTIVO",
                'Fecha Retiro': match.fecha_retiro or "No aplica"
            })

        response_data = dict(results[0])
        response_data['encontrado'] = True
        response_data['resultados'] = results
        return JsonResponse(response_data)
    except Exception as e:
        return JsonResponse({'encontrado': False, 'resultados': [], 'error': str(e)}, status=200)

from django.views.decorators.csrf import csrf_exempt
from django.db import transaction

@csrf_exempt
def api_sincronizar_afiliados(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Metodo no permitido'}, status=405)
    try:
        data = json.loads(request.body.decode('utf-8'))
        records = data.get('registros', [])
        secret = data.get('secret', '')
        if secret != 'nexus_cloud_sync_secret_2026':
            return JsonResponse({'error': 'No autorizado'}, status=401)
            
        if not records:
            return JsonResponse({'mensaje': 'No hay registros para sincronizar', 'sincronizados': 0})
            
        updated = 0
        with transaction.atomic():
            if data.get('clear'):
                Carpeta.objects.filter(categoria='TRABAJADOR').delete()
                
            for r in records:
                cedula = re.sub(r"\D", "", str(r.get('cedula') or r.get('Cédula') or ''))
                if not cedula:
                    continue
                    
                nombre = str(r.get('nombre') or r.get('Nombre') or 'SIN NOMBRE').upper().strip()
                fecha = str(r.get('fecha') or r.get('Fecha') or '')
                tipo_id = str(r.get('tipo_identificacion') or r.get('Tipo Identificación') or 'CEDULA')
                estado = str(r.get('estado') or r.get('Estado') or 'ACTIVO')
                fecha_retiro = str(r.get('fecha_retiro') or r.get('Fecha Retiro') or '')
                
                try:
                    modulo = int(r.get('modulo') or r.get('Módulo') or 1)
                    estante = int(r.get('estante') or r.get('Estante') or 1)
                    bandeja = int(r.get('bandeja') or r.get('Bandeja') or 1)
                    cubiculo = int(r.get('cubiculo') or r.get('Cubículo') or 1)
                    num_carpeta = int(r.get('numero_carpeta') or r.get('Número de Carpeta') or 1)
                except Exception:
                    modulo, estante, bandeja, cubiculo, num_carpeta = 1, 1, 1, 1, 1

                Carpeta.objects.update_or_create(
                    categoria='TRABAJADOR',
                    modulo=modulo,
                    estante=estante,
                    bandeja=bandeja,
                    cubiculo=cubiculo,
                    numero_carpeta=num_carpeta,
                    defaults={
                        'identificacion': cedula,
                        'nombre': nombre,
                        'fecha': fecha,
                        'tipo_identificacion': tipo_id,
                        'estado': estado,
                        'fecha_retiro': fecha_retiro,
                    }
                )
                updated += 1
                
        return JsonResponse({'exito': True, 'sincronizados': updated, 'total_db': Carpeta.objects.count()})
    except Exception as e:
        return JsonResponse({'exito': False, 'error': str(e)}, status=200)


# Vista del panel principal, requiere inicio de sesión
@login_required
def dashboard(request):
    # Solo Administrador (SUPER) y Jefe de Archivo (JEFE) tienen acceso al Dashboard Gerencial
    if not is_jefe(request.user):
        return redirect('gestion_documental')
        
    es_admin = is_super(request.user)
    historial = HistorialCarpeta.objects.all().order_by('-fecha')[:10]

    # Estadísticas para el Dashboard Interactivo
    total_carpetas = Carpeta.objects.count()
    activos_cnt = Carpeta.objects.filter(estado='ACTIVO').count()
    inactivos_cnt = Carpeta.objects.filter(estado='INACTIVO').count()
    muertos_cnt = Carpeta.objects.filter(estado='MUERTO').count()
    
    # Calcular Alertas Naranja (+10 años inactivos)
    hoy = datetime.date.today()
    diez_anos_atras = hoy.year - 10
    alertas_naranja = 0
    
    for c in Carpeta.objects.filter(estado='INACTIVO'):
        fr = str(c.fecha_retiro or '')
        years = re.findall(r'\b(19\d\d|20\d\d)\b', fr)
        if years:
            try:
                y = int(years[0])
                if y <= diez_anos_atras:
                    alertas_naranja += 1
            except Exception:
                pass

    return render(request, 'dashboard/index.html', {
        'total_carpetas': total_carpetas,
        'activos_cnt': activos_cnt,
        'inactivos_cnt': inactivos_cnt,
        'muertos_cnt': muertos_cnt,
        'alertas_naranja': alertas_naranja,
        'es_admin': es_admin,
        'historial': historial,
        'header_title': 'Panel Principal'
    })

# MÓDULOS DE VENTANILLA ÚNICA Y CORRESPONDENCIA DEBAJO ELIMINADOS

# === VISTAS DE GESTIÓN DOCUMENTAL Y ARCHIVO ===

# Vista principal para la administración del archivo físico
@login_required
def gestion_documental(request):
    # Obtiene la categoría actual de la URL (por defecto TRABAJADOR)
    cat_activa = request.GET.get('cat', 'TRABAJADOR')
    # Obtiene el término de búsqueda
    query = request.GET.get('q', '').strip()
    
    # Obtiene todas las carpetas de la categoría seleccionada, ordenadas por fecha de registro (últimas creadas primero)
    carpetas = Carpeta.objects.filter(categoria=cat_activa).order_by('-fecha_registro')
    
    if query:
        clean_q = re.sub(r"\D", "", query)
        if clean_q:
            carpetas = carpetas.filter(Q(identificacion__icontains=clean_q) | Q(nombre__icontains=query))
        else:
            carpetas = carpetas.filter(nombre__icontains=query)
            
    from django.core.paginator import Paginator
    paginator = Paginator(carpetas, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'afiliados/gestion_documental.html', {
        'carpetas': page_obj,
        'cat_activa': cat_activa,
        'query': query,
        'can_edit': is_aux_or_higher(request.user),
        'header_title': 'Gestión Documental'
    })

# Vista para ver los detalles de una carpeta física y su contenido
@login_required
def detalle_carpeta(request, carpeta_id):
    # Busca la carpeta por ID o por identificación/cédula
    it = Carpeta.objects.filter(id=carpeta_id).first()
    if not it:
        it = Carpeta.objects.filter(identificacion=str(carpeta_id)).first()
        
    if not it:
        messages.warning(request, f"El expediente solicitado (ID: {carpeta_id}) no fue encontrado o fue reubicado.")
        return redirect('gestion_documental')
        
    # Obtiene los documentos asociados
    documentos = it.documentos.all().order_by('-fecha_creacion')
    
    if request.method == 'POST':
        # LÓGICA DE EDICIÓN DE CARPETA
        if 'editar_carpeta' in request.POST:
            if not is_aux_or_higher(request.user):
                messages.error(request, "No tiene permisos para editar expedientes.")
                return redirect('detalle_carpeta', carpeta_id=it.id)
            form = CarpetaForm(request.POST, instance=it)
            if form.is_valid():
                try:
                    actualizada = form.save()
                    
                    # Registrar acción en historial
                    HistorialCarpeta.objects.create(
                        carpeta=it,
                        usuario=request.user,
                        accion='EDICION',
                        observaciones='Se actualizaron los datos del expediente.'
                    )
                    
                    messages.success(request, "Datos del expediente actualizados correctamente.")
                except Exception as e:
                    messages.error(request, f"Error al actualizar: {str(e)}")
            else:
                messages.error(request, "Error en el formulario. Verifique los datos.")
        
        # LÓGICA DE CARGA DE DOCUMENTOS
        elif 'subir_doc' in request.POST:
            if not is_aux_or_higher(request.user):
                messages.error(request, "No tiene permisos para digitalizar o subir documentos.")
                return redirect('detalle_carpeta', carpeta_id=it.id)
            nombre_doc = request.POST.get('nombre_doc')
            archivo = request.FILES.get('archivo')
            if nombre_doc and archivo:
                if not str(archivo.name).lower().endswith('.pdf'):
                    messages.error(request, "⚠️ Formato no permitido. Únicamente se pueden subir documentos en formato PDF.")
                    return redirect('detalle_carpeta', carpeta_id=it.id)
                    
                Documento.objects.create(
                    nombre=nombre_doc,
                    archivo=archivo,
                    carpeta=it
                )
                # Registrar carga de documento en historial
                HistorialCarpeta.objects.create(
                    carpeta=it,
                    usuario=request.user,
                    accion='DOCUMENTO',
                    observaciones=f'Se subió el archivo PDF: {nombre_doc}'
                )
                messages.success(request, f"Documento PDF '{nombre_doc}' subido con éxito a la bodega digital.")
            else:
                messages.error(request, "Debe proporcionar un nombre y un archivo PDF válido.")
        
        return redirect('detalle_carpeta', carpeta_id=it.id)
    
    form = CarpetaForm(instance=it)
    
    # Obtener historial de auditoría de esta carpeta específica (últimos 20 eventos)
    historial_carpeta = HistorialCarpeta.objects.filter(carpeta=it).select_related('usuario').order_by('-fecha')[:20]
    
    return render(request, 'afiliados/detalle.html', {
        'carpeta': it,
        'form': form,
        'documentos': documentos,
        'historial_carpeta': historial_carpeta,
        'can_edit': is_aux_or_higher(request.user),
        'can_delete_doc': is_super(request.user),
        'header_title': f'Expediente: {it.nombre}'
    })


# Vista para eliminar definitivamente un expediente fÃ­sico del sistema
@login_required
def borrar_carpeta(request, carpeta_id):
    if not is_super(request.user):
        messages.error(request, "No tiene permisos para eliminar expedientes del sistema.")
        return redirect('gestion_documental')
    # Busca la carpeta
    it = get_object_or_404(Carpeta, id=carpeta_id)
    # Guarda el nombre para mostrarlo en el mensaje
    nom = it.nombre
    # Elimina el registro de la base de datos
    it.delete()
    # Muestra un aviso de eliminaciÃ³n
    messages.warning(request, f"Expediente {nom} eliminado.")
    # Regresa al listado principal
    return redirect('gestion_documental')

@login_required
def borrar_documento(request, doc_id):
    if not is_super(request.user):
        from documentos.models import Documento
        it = get_object_or_404(Documento, id=doc_id)
        messages.error(request, "No tiene permisos para eliminar documentos digitales.")
        return redirect('detalle_carpeta', carpeta_id=it.carpeta.id)
        
    # Buscar el documento por su ID y borrarlo
    from documentos.models import Documento
    it = get_object_or_404(Documento, id=doc_id)
    carpeta_id = it.carpeta.id
    it.delete()
    messages.warning(request, "Documento PDF eliminado digitalmente.")
    return redirect('detalle_carpeta', carpeta_id=carpeta_id)

from django.http import HttpResponse
from openpyxl import Workbook

@login_required
def mapa_visual(request):
    nivel = 'CATEGORIA'
    cat = request.GET.get('cat')
    mod_id = request.GET.get('mod')
    est_id = request.GET.get('est')
    ban_id = request.GET.get('ban')
    cub_id = request.GET.get('cub')
    
    objetos = []
    breadcrumb = []
    
    if cat:
        nivel = 'MODULO'
        breadcrumb.append({'nombre': f'Archivo {cat}', 'url': f'?cat={cat}'})
        
        if mod_id:
            nivel = 'ESTANTE'
            breadcrumb.append({'nombre': f'MÃ³dulo {mod_id}', 'url': f'?cat={cat}&mod={mod_id}'})
            
            if est_id:
                nivel = 'BANDEJA'
                breadcrumb.append({'nombre': f'Estante {est_id}', 'url': f'?cat={cat}&mod={mod_id}&est={est_id}'})
                
                if ban_id:
                    nivel = 'CUBICULO'
                    breadcrumb.append({'nombre': f'Bandeja {ban_id}', 'url': f'?cat={cat}&mod={mod_id}&est={est_id}&ban={ban_id}'})
                    
                    if cub_id:
                        nivel = 'CARPETA'
                        breadcrumb.append({'nombre': f'CubÃ­culo {cub_id}', 'url': f'?cat={cat}&mod={mod_id}&est={est_id}&ban={ban_id}&cub={cub_id}'})
                        objetos = Carpeta.objects.filter(categoria=cat, modulo=mod_id, estante=est_id, bandeja=ban_id, cubiculo=cub_id).order_by('numero_carpeta')
                    else:
                        cub_nums = Carpeta.objects.filter(categoria=cat, modulo=mod_id, estante=est_id, bandeja=ban_id).values_list('cubiculo', flat=True).distinct().order_by('cubiculo')
                        objetos = [{'id': num, 'numero': num} for num in cub_nums]
                else:
                    ban_nums = Carpeta.objects.filter(categoria=cat, modulo=mod_id, estante=est_id).values_list('bandeja', flat=True).distinct().order_by('bandeja')
                    objetos = [{'id': num, 'numero': num} for num in ban_nums]
            else:
                est_nums = Carpeta.objects.filter(categoria=cat, modulo=mod_id).values_list('estante', flat=True).distinct().order_by('estante')
                objetos = [{'id': num, 'numero': num} for num in est_nums]
        else:
            mod_nums = Carpeta.objects.filter(categoria=cat).values_list('modulo', flat=True).distinct().order_by('modulo')
            objetos = [{'id': num, 'numero': num} for num in mod_nums]

    return render(request, 'afiliados/mapa_visual.html', {
        'nivel': nivel, 'objetos': objetos, 'breadcrumb': breadcrumb, 'cat': cat
    })

def calcular_anos_inactividad(fecha_retiro_str):
    if not fecha_retiro_str:
        return None
    fr = str(fecha_retiro_str).strip()
    if fr.lower() in ['nan', 'none', 'no aplica', '']:
        return None
    
    import re, datetime
    match = re.findall(r'\b(19\d\d|20\d\d)\b', fr)
    if match:
        ano = int(match[0])
        hoy = datetime.date.today()
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d', '%d-%m-%Y'):
            try:
                dt = datetime.datetime.strptime(fr[:10], fmt).date()
                diff_days = (hoy - dt).days
                return round(diff_days / 365.25, 1)
            except Exception:
                pass
        return float(max(0, hoy.year - ano))
    return None

def filtrar_carpetas_reporte(request):
    estado = request.GET.get('estado', 'TODOS').strip().upper()
    categoria = request.GET.get('categoria', 'TODOS').strip().upper()
    modulo = request.GET.get('modulo', '').strip()
    anos_min_str = request.GET.get('anos_min', '').strip()
    anos_max_str = request.GET.get('anos_max', '').strip()
    
    anos_min = float(anos_min_str) if anos_min_str else None
    anos_max = float(anos_max_str) if anos_max_str else None
    
    qs = Carpeta.objects.all()
    
    if categoria in ['TRABAJADOR', 'PATRONAL']:
        qs = qs.filter(categoria=categoria)
        
    if estado and estado not in ['TODOS', 'TODOS LOS ESTADOS', '']:
        qs = qs.filter(estado__iexact=estado)
        
    if modulo:
        try:
            qs = qs.filter(modulo=int(modulo))
        except ValueError:
            pass
            
    # Filtrar por rango de años de inactividad
    resultados = []
    for c in qs:
        anos_inactivo = calcular_anos_inactividad(c.fecha_retiro)
        
        # Si se especificó rango de años de inactividad
        if anos_min is not None or anos_max is not None:
            if anos_inactivo is None:
                continue
            if anos_min is not None and anos_inactivo < anos_min:
                continue
            if anos_max is not None and anos_inactivo > anos_max:
                continue
                
        resultados.append({
            'carpeta': c,
            'anos_inactivo': anos_inactivo
        })

    return resultados, {
        'estado': estado,
        'categoria': categoria,
        'modulo': modulo,
        'anos_min': anos_min_str,
        'anos_max': anos_max_str
    }

@login_required
def panel_reportes(request):
    if not is_jefe(request.user):
        messages.error(request, "Solo el Administrador y el Jefe de Archivo tienen autorización para acceder al Generador de Reportes.")
        return redirect('gestion_documental')
        
    resultados, params = filtrar_carpetas_reporte(request)
    
    # Módulos disponibles para el selector
    modulos_disp = sorted(list(Carpeta.objects.values_list('modulo', flat=True).distinct()))
    
    from django.core.paginator import Paginator
    paginator = Paginator(resultados, 25)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'afiliados/reportes.html', {
        'resultados_page': page_obj,
        'total_encontrados': len(resultados),
        'params': params,
        'modulos_disp': modulos_disp,
        'header_title': 'Generador de Reportes Parametrizados'
    })

@login_required
def exportar_excel_archivo(request):
    if not is_jefe(request.user):
        messages.error(request, "No tiene permisos para exportar reportes institucionales.")
        return redirect('gestion_documental')
        
    resultados, params = filtrar_carpetas_reporte(request)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Filtrado"
    
    # Encabezados
    headers = [
        'CATEGORÍA', 'TIPO ID', 'IDENTIFICACIÓN', 'NOMBRE COMPLETO', 
        'ESTADO', 'FECHA RETIRO', 'AÑOS INACTIVO', 
        'MÓDULO', 'ESTANTE', 'BANDEJA', 'CUBÍCULO', 'CARPETA #'
    ]
    ws.append(headers)
    
    # Estilos encabezado
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_fill = PatternFill(start_color="004A87", end_color="004A87", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        
    for item in resultados:
        c = item['carpeta']
        anos_inact = f"{item['anos_inactivo']} años" if item['anos_inactivo'] is not None else "No aplica"
        ws.append([
            c.get_categoria_display(),
            c.tipo_identificacion or "CC",
            c.identificacion,
            c.nombre,
            c.estado or "ACTIVO",
            c.fecha_retiro or "No aplica",
            anos_inact,
            c.modulo,
            c.estante,
            c.bandeja,
            c.cubiculo,
            c.numero_carpeta
        ])
        
    # Autoajuste de columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
    
    nombre_archivo = f"Reporte_Nexus_{params['estado']}_{params['categoria']}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response

@login_required
def historial_auditoria(request):
    if not is_jefe(request.user):
        messages.error(request, "No tiene autorización para ver el historial de auditoría.")
        return redirect('gestion_documental')

    import re
    from django.db import connection
    from django.core.paginator import Paginator
    
    # 1. Obtener registros de la base de datos (Plataforma Web)
    web_records = []
    for h in HistorialCarpeta.objects.all().select_related('usuario', 'carpeta').order_by('-fecha'):
        web_records.append({
            'fecha': h.fecha,
            'usuario': h.usuario.username,
            'accion': h.accion,
            'detalles': f"Expediente: {h.carpeta.nombre} ({h.carpeta.identificacion}). {h.observaciones or ''}",
            'plataforma': 'Web'
        })
        
    # 2. Obtener registros del archivo de texto (Plataforma Escritorio)
    desktop_records = []
    try:
        db_file = connection.settings_dict['NAME']
        log_file = os.path.join(os.path.dirname(db_file), "historial_cambios.txt")
        if os.path.exists(log_file):
            pattern = re.compile(r"^\[(.*?)\]\s+\[Usuario:\s*(.*?)\]\s+(.*?):\s*(.*)$")
            with open(log_file, "r", encoding="utf-8", errors="ignore") as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    match = pattern.match(line)
                    if match:
                        timestamp_str, user, action, details = match.groups()
                        try:
                            dt = datetime.datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S")
                            dt = timezone.make_aware(dt) if timezone.is_naive(dt) else dt
                        except Exception:
                            dt = timezone.now()
                        desktop_records.append({
                            'fecha': dt,
                            'usuario': user,
                            'accion': action,
                            'detalles': details,
                            'plataforma': 'Escritorio'
                        })
    except Exception as e:
        print(f"Error al leer historial_cambios.txt: {e}")
        
    # 3. Combinar y ordenar por fecha descendente
    all_records = web_records + desktop_records
    all_records.sort(key=lambda x: x['fecha'], reverse=True)
    
    # 4. Aplicar filtros
    plat_filter = request.GET.get('plataforma', 'Todos')
    user_filter = request.GET.get('usuario', '').strip().lower()
    action_filter = request.GET.get('accion', 'Todos')
    q_filter = request.GET.get('q', '').strip().lower()
    
    filtered_records = []
    for r in all_records:
        if plat_filter != 'Todos' and r['plataforma'] != plat_filter:
            continue
        if action_filter != 'Todos' and r['accion'].upper() != action_filter.upper():
            continue
        if user_filter and user_filter not in r['usuario'].lower():
            continue
        if q_filter and q_filter not in r['detalles'].lower():
            continue
        filtered_records.append(r)
        
    # 5. Paginación
    paginator = Paginator(filtered_records, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # 6. Consultar Sesiones Activas en Tiempo Real
    from .models import ActiveUserSession, LoginLog
    # Limpiar sesiones con más de 24 horas de inactividad
    un_dia_atras = timezone.now() - datetime.timedelta(hours=24)
    ActiveUserSession.objects.filter(last_activity__lt=un_dia_atras).delete()
    active_sessions = ActiveUserSession.objects.all().select_related('user', 'user__profile').order_by('-last_activity')
    
    # 7. Consultar Historial de Inicios de Sesión
    login_logs = LoginLog.objects.all().select_related('user').order_by('-timestamp')[:50]
    
    tab_activa = request.GET.get('tab', 'sesiones' if is_super(request.user) else 'trazabilidad')
    
    return render(request, 'afiliados/historial_auditoria.html', {
        'page_obj': page_obj,
        'plat_filter': plat_filter,
        'user_filter': user_filter,
        'action_filter': action_filter,
        'q_filter': q_filter,
        'total_eventos': len(filtered_records),
        'active_sessions': active_sessions,
        'login_logs': login_logs,
        'tab_activa': tab_activa,
        'is_super': is_super(request.user),
        'header_title': 'Centro de Seguridad y Auditoría'
    })

@login_required
def cerrar_sesion_remota(request, session_id):
    if not is_super(request.user):
        messages.error(request, "No tiene permisos para cerrar sesiones de otros usuarios.")
        return redirect('historial_auditoria')
    
    from django.contrib.sessions.models import Session
    from .models import ActiveUserSession
    try:
        active_sess = ActiveUserSession.objects.filter(id=session_id).first()
        if active_sess:
            username = active_sess.user.username
            Session.objects.filter(session_key=active_sess.session_key).delete()
            active_sess.delete()
            messages.success(request, f"Sesión del usuario '{username}' cerrada remotamente.")
        else:
            messages.warning(request, "La sesión ya no se encuentra activa.")
    except Exception as e:
        messages.error(request, f"Error al cerrar la sesión: {str(e)}")
        
    return redirect('/historial-auditoria/?tab=sesiones')
def exportar_auditoria_excel(request):
    if not is_super(request.user):
        messages.error(request, "No tiene autorización para exportar auditorías.")
        return redirect('dashboard')

    import re
    from django.db import connection
    
    # 1. Obtener registros de la base de datos (Plataforma Web)
    web_records = []
    for h in HistorialCarpeta.objects.all().select_related('usuario', 'carpeta').order_by('-fecha'):
        web_records.append({
            'fecha': h.fecha,
            'usuario': h.usuario.username,
            'accion': h.accion,
            'detalles': f"Expediente: {h.carpeta.nombre} ({h.carpeta.identificacion}). {h.observaciones or ''}",
            'plataforma': 'Web'
        })
        
    # 2. Obtener registros del archivo de texto (Plataforma Escritorio)
    desktop_records = []
    try:
        db_file = connection.settings_dict['NAME']
        log_file = os.path.join(os.path.dirname(db_file), "historial_cambios.txt")
        if os.path.exists(log_file):
            pattern = re.compile(r"^\[(.*?)\]\s+\[Usuario:\s*(.*?)\]\s+(.*?):\s*(.*)$")
            with open(log_file, "r", encoding="utf-8", errors="ignore") as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    match = pattern.match(line)
                    if match:
                        timestamp_str, user, action, details = match.groups()
                        try:
                            dt = datetime.datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S")
                            dt = timezone.make_aware(dt) if timezone.is_naive(dt) else dt
                        except Exception:
                            dt = timezone.now()
                        desktop_records.append({
                            'fecha': dt,
                            'usuario': user,
                            'accion': action,
                            'detalles': details,
                            'plataforma': 'Escritorio'
                        })
    except Exception as e:
        print(f"Error al leer historial_cambios.txt: {e}")
        
    # 3. Combinar y ordenar por fecha descendente
    all_records = web_records + desktop_records
    all_records.sort(key=lambda x: x['fecha'], reverse=True)
    
    # 4. Aplicar filtros
    plat_filter = request.GET.get('plataforma', 'Todos')
    user_filter = request.GET.get('usuario', '').strip().lower()
    action_filter = request.GET.get('accion', 'Todos')
    q_filter = request.GET.get('q', '').strip().lower()
    
    filtered_records = []
    for r in all_records:
        if plat_filter != 'Todos' and r['plataforma'] != plat_filter:
            continue
        if action_filter != 'Todos' and r['accion'].upper() != action_filter.upper():
            continue
        if user_filter and user_filter not in r['usuario'].lower():
            continue
        if q_filter and q_filter not in r['detalles'].lower():
            continue
        filtered_records.append(r)
        
    # 5. Generar Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Historial Auditoria Global"
    
    headers = ['FECHA/HORA', 'PLATAFORMA', 'USUARIO', 'ACCIÓN', 'DETALLES DE EVENTO']
    ws.append(headers)
    
    for r in filtered_records:
        fecha_str = r['fecha'].strftime('%Y-%m-%d %H:%M:%S')
        ws.append([
            fecha_str,
            r['plataforma'],
            r['usuario'],
            r['accion'],
            r['detalles']
        ])
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Reporte_Auditoria_Global.xlsx'
    wb.save(response)
    return response

# FIN DE VISTAS DE AFILIADOS

